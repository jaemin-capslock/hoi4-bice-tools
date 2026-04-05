#!/usr/bin/env python3
"""
bice_viz.py — Excel visualisation for BICE division stats.

Public API
----------
generate_excel(equip_db, bat_db, div_stats, output_path)
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bice_parser import LAND_EQUIP_FAMILIES, LAND_FAMILY_SET


# ─────────────────────────────────────────────
#  COLOUR PALETTE
# ─────────────────────────────────────────────

C_HEADER_DARK  = "1F3864"
C_HEADER_MED   = "2E75B6"
C_HEADER_LIGHT = "BDD7EE"
C_ALT_ROW      = "EBF3FB"
C_GOLD         = "C9A84C"
C_WHITE        = "FFFFFF"
C_RED_LIGHT    = "FCE4D6"
C_GREEN_LIGHT  = "E2EFDA"
C_ORANGE_LIGHT = "FFF2CC"

_THIN   = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─────────────────────────────────────────────
#  CELL HELPERS
# ─────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value, *,
         bold=True, bg=C_HEADER_DARK, fg=C_WHITE,
         size=11, wrap=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap)
    cell.border    = _BORDER
    return cell


def _cell(ws, row: int, col: int, value, *,
          bg=None, bold=False, align="center", fmt=None, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=size)
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _auto_width(ws, min_w=8, max_w=40):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max(length + 2, min_w), max_w)


# ─────────────────────────────────────────────
#  SHEET 1: DIVISION COMPARISON
# ─────────────────────────────────────────────

_DIV_COLS = [
    "Name", "Width", "HP", "Org", "Soft Attack", "Hard Attack",
    "Air Attack", "Defense", "Breakthrough", "Hardness",
    "Collateral", "Suppression", "Speed (km/h)",
    "Manpower", "Weight", "Supply/day", "IC Cost",
    "Training (days)", "Notes",
]

_COMBAT_STAT_COLS = {"Soft Attack", "Hard Attack", "Defense", "Breakthrough"}


def write_divisions_sheet(wb: Workbook, div_stats: list[dict]) -> None:
    ws = wb.create_sheet("Division Comparison")
    ws.freeze_panes = "B3"

    ncols = len(_DIV_COLS)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws.cell(row=1, column=1, value="BICE Division Templates — Calculated Stats")
    t.font      = Font(bold=True, size=14, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, label in enumerate(_DIV_COLS, 1):
        _hdr(ws, 2, ci, label, bg=C_HEADER_MED, wrap=True)

    alternating = [C_WHITE, C_ALT_ROW]
    for ri, stats in enumerate(div_stats, 3):
        bg = alternating[(ri - 3) % 2]
        for ci, key in enumerate(_DIV_COLS, 1):
            # Map sheet column name → stats dict key
            val = stats.get(key, stats.get("name" if key == "Name" else
                            "notes" if key == "Notes" else key, ""))
            is_float = isinstance(val, float)
            _cell(ws, ri, ci, val, bg=bg,
                  bold=(ci == 1),
                  align="left" if ci in (1, ncols) else "center",
                  fmt="#,##0.00" if is_float else None)

    # Heat-colour key combat stats (green=high, red=low)
    if div_stats:
        for key in _COMBAT_STAT_COLS:
            ci = _DIV_COLS.index(key) + 1
            vals = [s.get(key, 0) for s in div_stats
                    if isinstance(s.get(key), (int, float))]
            if not vals:
                continue
            lo, hi = min(vals), max(vals)
            for ri, stats in enumerate(div_stats, 3):
                v = stats.get(key, 0)
                if isinstance(v, (int, float)) and hi > lo:
                    t = (v - lo) / (hi - lo)
                    r = int(255 - t * 80)
                    g = int(175 + t * 80)
                    b = 175
                    ws.cell(row=ri, column=ci).fill = PatternFill(
                        "solid", fgColor=f"{r:02X}{g:02X}{b:02X}")

    _auto_width(ws)


# ─────────────────────────────────────────────
#  SHEET 2: EQUIPMENT REFERENCE
# ─────────────────────────────────────────────

_EQUIP_COLS = [
    ("ID",           "id"),
    ("Family",       "family_label"),
    ("Year",         "year"),
    ("SA",           "soft_attack"),
    ("HA",           "hard_attack"),
    ("AA",           "air_attack"),
    ("Defense",      "defense"),
    ("BT",           "breakthrough"),
    ("AP",           "ap_attack"),
    ("Armor",        "armor_value"),
    ("Hardness",     "hardness"),
    ("Reliability",  "reliability"),
    ("Speed",        "maximum_speed"),
    ("IC Cost",      "build_cost_ic"),
    ("Collateral",   "additional_collateral_damage"),
    ("Suppression",  "suppression"),
    ("Fuel Use",     "fuel_consumption"),
]

_PALETTE = [
    C_ALT_ROW, C_GREEN_LIGHT, C_ORANGE_LIGHT, "E8D5F5", "D5F5E3",
    "FAD7A0", "AED6F1", "F9E79F", "D5DBDB", "F8BBD0",
]


def write_equipment_sheet(wb: Workbook, equip_db: dict) -> None:
    ws = wb.create_sheet("Equipment Reference")
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    ncols = len(_EQUIP_COLS)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws.cell(row=1, column=1, value="BICE Equipment Reference — Land Equipment")
    t.font      = Font(bold=True, size=14, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")

    for ci, (label, _) in enumerate(_EQUIP_COLS, 1):
        _hdr(ws, 2, ci, label, bg=C_HEADER_MED)

    land = {
        eid: eq for eid, eq in equip_db.items()
        if eq.get("family") in LAND_FAMILY_SET
        and eq.get("is_archetype") not in (True, "yes", 1)
        and eq.get("active") not in (False, "no")
    }
    rows = sorted(land.values(),
                  key=lambda e: (e.get("family_label", ""), e.get("year", 0)))

    family_colours: dict[str, str] = {}
    colour_idx = 0
    for ri, eq in enumerate(rows, 3):
        fam = eq.get("family_label", "")
        if fam not in family_colours:
            family_colours[fam] = _PALETTE[colour_idx % len(_PALETTE)]
            colour_idx += 1
        bg = family_colours[fam]
        for ci, (_, key) in enumerate(_EQUIP_COLS, 1):
            val = eq.get(key, "")
            if isinstance(val, float) and val == 0.0:
                val = ""
            _cell(ws, ri, ci, val, bg=bg,
                  align="left" if ci == 1 else "center",
                  fmt="#,##0.00" if isinstance(val, float) else None)

    _auto_width(ws)


# ─────────────────────────────────────────────
#  SHEET 3: BATTALION REFERENCE
# ─────────────────────────────────────────────

_BAT_COLS = [
    ("Unit ID",        "id"),
    ("Source File",    "source"),
    ("Group",          "group"),
    ("HP",             "max_strength"),
    ("Base Org",       "max_organisation"),
    ("Combat Width",   "combat_width"),
    ("Manpower",       "manpower"),
    ("Weight",         "weight"),
    ("Supply/day",     "supply_consumption"),
    ("Training",       "training_time"),
    ("Base BT mod",    "breakthrough"),
    ("Base SA mod",    "soft_attack"),
    ("Base Def mod",   "defense"),
    ("Suppression",    "suppression"),
    ("Transport",      "transport"),
    ("Equipment Need", "_need"),
]

_LAND_GROUPS = {
    "infantry", "artillery", "armor", "motorized",
    "mechanized", "anti_tank", "anti_air", "support",
}

_GROUP_COLOURS = {
    "infantry":  C_GREEN_LIGHT,
    "artillery": C_ORANGE_LIGHT,
    "armor":     C_RED_LIGHT,
    "motorized": "DEEBF7",
    "support":   "EDE7F6",
}


def write_battalion_sheet(wb: Workbook, bat_db: dict) -> None:
    ws = wb.create_sheet("Battalion Reference")
    ws.freeze_panes = "A3"

    ncols = len(_BAT_COLS)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws.cell(row=1, column=1, value="BICE Battalion / Sub-Unit Reference")
    t.font      = Font(bold=True, size=14, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, (label, _) in enumerate(_BAT_COLS, 1):
        _hdr(ws, 2, ci, label, bg=C_HEADER_MED, wrap=True)

    land_bats = {
        uid: b for uid, b in bat_db.items()
        if b.get("group", "") in _LAND_GROUPS or
           any(g in _LAND_GROUPS for g in b.get("categories", []))
    }
    rows = sorted(land_bats.values(),
                  key=lambda b: (b.get("group", ""), b.get("id", "")))

    for ri, bat in enumerate(rows, 3):
        bg = _GROUP_COLOURS.get(bat.get("group", ""), C_ALT_ROW)
        for ci, (_, key) in enumerate(_BAT_COLS, 1):
            if key == "_need":
                val = " | ".join(f"{k}×{v}" for k, v in bat.get("need", {}).items())
            else:
                val = bat.get(key, "")
            if isinstance(val, list):  # duplicate HOI4 keys → take last
                val = val[-1] if val else ""
            if val == 0.0 and isinstance(val, float):
                val = ""
            _cell(ws, ri, ci, val, bg=bg,
                  align="left" if ci in (1, 2, 15, 16) else "center")

    _auto_width(ws)


# ─────────────────────────────────────────────
#  SHEET 4: RAW EQUIPMENT DUMP
# ─────────────────────────────────────────────

def write_raw_dump_sheet(wb: Workbook, equip_db: dict) -> None:
    """Dump all equipment (including non-land) for debugging."""
    ws = wb.create_sheet("All Equipment (raw)")
    all_keys = sorted({k for eq in equip_db.values() for k in eq})

    for ci, k in enumerate(all_keys, 1):
        _hdr(ws, 1, ci, k, bg=C_HEADER_MED, size=9)

    for ri, eq in enumerate(
        sorted(equip_db.values(), key=lambda e: e.get("id", "")), 2
    ):
        for ci, k in enumerate(all_keys, 1):
            v = eq.get(k, "")
            # openpyxl can't serialise dicts or lists
            if isinstance(v, (dict, list)):
                v = str(v)
            ws.cell(row=ri, column=ci, value=v)

    _auto_width(ws, min_w=6, max_w=30)


# ─────────────────────────────────────────────
#  TOP-LEVEL ENTRY POINT
# ─────────────────────────────────────────────

def generate_excel(
    equip_db: dict,
    bat_db: dict,
    div_stats: list[dict],
    output_path: Path | str | None = None,
) -> Path:
    """
    Build and save the BICE stats workbook.

    Parameters
    ----------
    equip_db    : from bice_parser.build_equipment_db()
    bat_db      : from bice_parser.build_battalion_db()
    div_stats   : list of division stat dicts from bice_calc.calc_division()
    output_path : where to save (default: ~/Desktop/BICE_Stats.xlsx)

    Returns
    -------
    The resolved output Path.
    """
    if output_path is None:
        output_path = Path(os.path.expanduser("~")) / "Desktop" / "BICE_Stats.xlsx"
    output_path = Path(output_path)

    wb = Workbook()
    wb.remove(wb.active)  # discard default empty sheet

    write_divisions_sheet(wb, div_stats)
    write_equipment_sheet(wb, equip_db)
    write_battalion_sheet(wb, bat_db)
    write_raw_dump_sheet(wb, equip_db)

    wb.save(output_path)
    return output_path
